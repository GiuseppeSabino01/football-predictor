from io import BytesIO

from openpyxl import load_workbook

from fantasy.catalog import make_player
from fantasy.export import build_listone_excel, restore_listone_excel
from fantasy.service import (
    auction_managers,
    auction_player_assignment,
    auction_player_tier,
    create_league,
    new_workspace,
    record_auction_purchase,
    rename_auction_manager,
    update_auction_assignments,
    update_list_assignments,
)


def test_listone_excel_contains_personal_tiers_and_roster_data() -> None:
    workspace = new_workspace()
    league = create_league(
        workspace,
        "Export listone",
        initial_budget=250,
        participants=None,
        game_mode="list",
        roster_slots={"P": 0, "D": 1, "C": 0, "A": 0},
    )
    player = make_player(name="Dimarco", team="INT", role="D", quote=32)
    player.update({"bonus": 0.72, "potential": 81, "profile": "Esterno offensivo"})
    tier = next(item for item in league["auction_tiers"] if item["name"] == "Semi-top")
    update_list_assignments(
        league,
        [{"player": player, "in_roster": True, "tier_id": tier["id"]}],
    )

    workbook = load_workbook(BytesIO(build_listone_excel([player], league)))
    sheet = workbook["Listone"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=True))

    assert workbook.sheetnames == ["Listone", "Fasce personali"]
    assert values["Giocatore"] == "Dimarco"
    assert values["Fascia personale"] == "Semi-top"
    assert values["In rosa"] == "Sì"
    assert values["Crediti"] == 32
    assert values["Propensione bonus"] == 72
    assert values["Potenziale"] == 81
    assert values["Profilo"] == "Esterno offensivo"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == sheet.dimensions


def test_auction_excel_contains_manager_and_paid_credits() -> None:
    workspace = new_workspace()
    league = create_league(
        workspace,
        "Export asta",
        initial_budget=500,
        participants=2,
        roster_slots={"P": 1, "D": 0, "C": 0, "A": 0},
    )
    player = make_player(name="Sommer", team="INT", role="P", quote=18)
    rival = auction_managers(league)[1]
    record_auction_purchase(league, str(rival["id"]), player, 27)

    workbook = load_workbook(BytesIO(build_listone_excel([player], league)))
    sheet = workbook["Listone"]
    headers = [cell.value for cell in sheet[1]]
    values = dict(zip(headers, [cell.value for cell in sheet[2]], strict=True))

    assert values["Fantaallenatore"] == rival["name"]
    assert values["Crediti"] == 27
    assert values["In rosa"] == "No"


def test_auction_excel_roundtrip_restores_tiers_managers_and_prices() -> None:
    source_workspace = new_workspace()
    source = create_league(
        source_workspace,
        "Backup asta",
        initial_budget=500,
        participants=3,
        roster_slots={"P": 1, "D": 1, "C": 0, "A": 0},
    )
    players = [
        make_player(name="Sommer", team="INT", role="P", quote=18),
        make_player(name="Dimarco", team="INT", role="D", quote=32),
        make_player(name="Di Lorenzo", team="NAP", role="D", quote=24),
    ]
    source_managers = auction_managers(source)
    rename_auction_manager(source, str(source_managers[1]["id"]), "Mario")
    record_auction_purchase(source, str(source_managers[0]["id"]), players[0], 41)
    record_auction_purchase(source, str(source_managers[1]["id"]), players[1], 27)
    tier = next(item for item in source["auction_tiers"] if item["name"] == "Semi-top")
    update_auction_assignments(
        source,
        [{"player": players[1], "update_assignment": False, "tier_id": tier["id"]}],
    )
    raw = build_listone_excel(players, source)

    target_workspace = new_workspace()
    target = create_league(
        target_workspace,
        "Asta recuperata",
        initial_budget=500,
        participants=3,
        roster_slots={"P": 1, "D": 1, "C": 0, "A": 0},
    )
    target_managers = auction_managers(target)
    record_auction_purchase(target, str(target_managers[2]["id"]), players[2], 11)

    result = restore_listone_excel(raw, players, target)

    sommer = auction_player_assignment(target, str(players[0]["id"]))
    dimarco = auction_player_assignment(target, str(players[1]["id"]))
    assert result == {
        "matched": 3,
        "purchases": 2,
        "tier_assignments": 1,
        "unmatched": [],
    }
    assert sommer and sommer["is_user"] is True
    assert sommer["purchase"]["price"] == 41
    assert dimarco and dimarco["manager_name"] == "Mario"
    assert dimarco["purchase"]["price"] == 27
    assert auction_player_assignment(target, str(players[2]["id"])) is None
    assert auction_player_tier(target, str(players[1]["id"]))["name"] == "Semi-top"


def test_list_mode_excel_roundtrip_supports_legacy_files_without_player_id() -> None:
    workspace = new_workspace()
    source = create_league(
        workspace,
        "Backup listone",
        initial_budget=250,
        participants=None,
        game_mode="list",
        roster_slots={"P": 0, "D": 1, "C": 0, "A": 0},
    )
    player = make_player(name="Dimarco", team="INT", role="D", quote=32)
    tier = next(item for item in source["auction_tiers"] if item["name"] == "Top")
    update_list_assignments(
        source,
        [{"player": player, "in_roster": True, "tier_id": tier["id"]}],
    )
    workbook = load_workbook(BytesIO(build_listone_excel([player], source)))
    sheet = workbook["Listone"]
    id_column = next(
        cell.column for cell in sheet[1] if cell.value == "ID giocatore"
    )
    sheet.delete_cols(id_column)
    legacy_output = BytesIO()
    workbook.save(legacy_output)

    target_workspace = new_workspace()
    target = create_league(
        target_workspace,
        "Listone recuperato",
        initial_budget=250,
        participants=None,
        game_mode="list",
        roster_slots={"P": 0, "D": 1, "C": 0, "A": 0},
    )

    result = restore_listone_excel(legacy_output.getvalue(), [player], target)

    assert result["purchases"] == 1
    assert target["purchases"][0]["player_id"] == player["id"]
    assert auction_player_tier(target, str(player["id"]))["name"] == "Top"
