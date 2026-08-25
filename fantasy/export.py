from __future__ import annotations

import math
import unicodedata
from copy import deepcopy
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from fantasy.catalog import catalog_dataframe
from fantasy.service import (
    GAME_MODE_AUCTION,
    GAME_MODE_LIST,
    auction_managers,
    auction_player_assignment,
    auction_player_tier,
    create_auction_tier,
    player_note,
    update_auction_assignments,
    update_list_assignments,
)

TIER_FILLS = {
    "red": "F4CCCC",
    "orange": "FCE5CD",
    "yellow": "FFF2CC",
    "green": "D9EAD3",
    "blue": "CFE2F3",
    "purple": "D9D2E9",
    "gray": "D9D9D9",
}
MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
RESTORE_HEADERS = {
    "Giocatore",
    "Squadra",
    "Ruolo",
    "Fascia personale",
    "In rosa",
    "Fantaallenatore",
    "Crediti",
}


def build_listone_excel(catalog: list[dict[str, Any]], league: dict[str, Any]) -> bytes:
    """Export the current player board with personal tiers and ownership data."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Listone"

    base_frame = catalog_dataframe(catalog)
    base_headers = [column for column in base_frame.columns if column != "_id"]
    extra_headers = ["Propensione bonus", "Potenziale", "Profilo"]
    headers = [
        "Giocatore",
        "ID giocatore",
        "Squadra",
        "Ruolo",
        "Fascia personale",
        "Note personali",
        "In rosa",
        "Fantaallenatore",
        "Crediti",
        *[
            header
            for header in base_headers
            if header not in {"Giocatore", "Squadra", "Ruolo"}
        ],
        *extra_headers,
    ]
    sheet.append(headers)

    roster_ids = {
        str(purchase.get("player_id")) for purchase in league.get("purchases", [])
    }
    tier_color_by_name = {
        str(tier.get("name") or ""): str(tier.get("color") or "gray")
        for tier in league.get("auction_tiers", [])
    }
    tier_counts = {name: 0 for name in tier_color_by_name}
    for player, (_, base_row) in zip(catalog, base_frame.iterrows(), strict=True):
        player_id = str(player.get("id") or "")
        personal_tier = auction_player_tier(league, player_id)
        personal_tier_name = (
            str(personal_tier.get("name") or "") if personal_tier else ""
        )
        if personal_tier_name:
            tier_counts[personal_tier_name] = tier_counts.get(personal_tier_name, 0) + 1

        assignment = (
            auction_player_assignment(league, player_id)
            if league.get("game_mode") == GAME_MODE_AUCTION
            else None
        )
        if assignment:
            in_roster = "Sì" if assignment.get("is_user") else "No"
            manager_name = str(assignment.get("manager_name") or "")
            credits = assignment.get("purchase", {}).get("price")
        else:
            in_roster = "Sì" if player_id in roster_ids else "No"
            manager_name = "La mia squadra" if player_id in roster_ids else ""
            credits = next(
                (
                    purchase.get("price")
                    for purchase in league.get("purchases", [])
                    if str(purchase.get("player_id")) == player_id
                ),
                None,
            )

        values = {
            "Giocatore": base_row.get("Giocatore"),
            "ID giocatore": player_id,
            "Squadra": base_row.get("Squadra"),
            "Ruolo": base_row.get("Ruolo"),
            "Fascia personale": personal_tier_name,
            "Note personali": player_note(league, player_id),
            "In rosa": in_roster,
            "Fantaallenatore": manager_name,
            "Crediti": credits,
            **{header: base_row.get(header) for header in base_headers},
            "Propensione bonus": _percentage_metric(player.get("bonus")),
            "Potenziale": _percentage_metric(player.get("potential")),
            "Profilo": player.get("profile"),
        }
        sheet.append([_excel_value(values.get(header)) for header in headers])
        if personal_tier_name:
            fill_color = TIER_FILLS.get(
                tier_color_by_name.get(personal_tier_name, "gray"), "D9D9D9"
            )
            for cell in sheet[sheet.max_row]:
                cell.fill = PatternFill("solid", fgColor=fill_color)

    _style_listone_sheet(sheet)
    id_column = headers.index("ID giocatore") + 1
    sheet.column_dimensions[get_column_letter(id_column)].hidden = True
    if sheet.max_row > 1:
        table = Table(displayName="ListoneFantacalcio", ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    tiers_sheet = workbook.create_sheet("Fasce personali")
    tiers_sheet.append(["Fascia", "Colore", "Giocatori assegnati"])
    for tier in league.get("auction_tiers", []):
        name = str(tier.get("name") or "")
        color = str(tier.get("color") or "gray")
        tiers_sheet.append([name, color, tier_counts.get(name, 0)])
        fill_color = TIER_FILLS.get(color, "D9D9D9")
        for cell in tiers_sheet[tiers_sheet.max_row]:
            cell.fill = PatternFill("solid", fgColor=fill_color)
    _style_tiers_sheet(tiers_sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def restore_listone_excel(
    raw: bytes,
    catalog: list[dict[str, Any]],
    league: dict[str, Any],
) -> dict[str, Any]:
    """Restore tiers and ownership from an Excel previously exported by the app."""
    if not raw:
        raise ValueError("Il file Excel e vuoto.")
    if len(raw) > MAX_IMPORT_BYTES:
        raise ValueError("Il file Excel supera il limite di 20 MB.")
    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(
            "Il file non e un Excel valido o risulta danneggiato."
        ) from exc
    if "Listone" not in workbook.sheetnames:
        raise ValueError(
            "Foglio 'Listone' non trovato: carica un Excel esportato dall'app."
        )

    sheet = workbook["Listone"]
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_cell_text(value) for value in raw_headers]
    missing_headers = sorted(RESTORE_HEADERS.difference(headers))
    if missing_headers:
        raise ValueError(
            "Colonne necessarie mancanti: " + ", ".join(missing_headers) + "."
        )

    rows: list[dict[str, Any]] = []
    for position, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if position > MAX_IMPORT_ROWS + 1:
            raise ValueError(
                "Il file contiene troppe righe per essere un listone valido."
            )
        row = dict(zip(headers, values, strict=False))
        if _cell_text(row.get("Giocatore")):
            rows.append(row)
    if not rows:
        raise ValueError("Il foglio 'Listone' non contiene giocatori.")

    matched_rows, unmatched = _match_restore_rows(rows, catalog)
    if not matched_rows:
        raise ValueError("Nessun giocatore del file corrisponde al listone attuale.")

    draft = deepcopy(league)
    tier_definitions = _tier_definitions(workbook, matched_rows, draft)
    draft["auction_tiers"] = []
    draft["auction_player_tiers"] = {}
    draft["auction_tiers_initialized"] = True
    restores_notes = "Note personali" in headers
    if restores_notes:
        draft["player_notes"] = {}
    tier_ids: dict[str, str] = {}
    for tier_name, color in tier_definitions:
        tier = create_auction_tier(draft, tier_name, color)
        tier_ids[_restore_key(tier_name)] = str(tier["id"])

    if draft.get("game_mode") == GAME_MODE_AUCTION:
        restored_purchases = _restore_auction_rows(draft, matched_rows, tier_ids)
    elif draft.get("game_mode") == GAME_MODE_LIST:
        restored_purchases = _restore_list_rows(draft, matched_rows, tier_ids)
    else:
        raise ValueError("Modalita del fantacalcio non riconosciuta.")

    league.clear()
    league.update(draft)
    return {
        "matched": len(matched_rows),
        "purchases": restored_purchases,
        "tier_assignments": sum(
            1 for row in matched_rows if _cell_text(row.get("Fascia personale"))
        ),
        "notes": sum(
            1 for row in matched_rows if _cell_text(row.get("Note personali"))
        ) if restores_notes else 0,
        "unmatched": unmatched,
    }


def _match_restore_rows(
    rows: list[dict[str, Any]], catalog: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[str]]:
    by_id = {str(player.get("id") or ""): player for player in catalog}
    by_full_key: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    by_name_role: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for player in catalog:
        name = _restore_key(player.get("name"))
        team = _restore_key(player.get("team"))
        role = _restore_key(player.get("role"))
        by_full_key.setdefault((name, team, role), []).append(player)
        by_name_role.setdefault((name, role), []).append(player)

    matched: list[dict[str, Any]] = []
    unmatched: list[str] = []
    used_player_ids: set[str] = set()
    for row in rows:
        player_id = _cell_text(row.get("ID giocatore"))
        player = by_id.get(player_id) if player_id else None
        if player is None:
            full_candidates = by_full_key.get(
                (
                    _restore_key(row.get("Giocatore")),
                    _restore_key(row.get("Squadra")),
                    _restore_key(row.get("Ruolo")),
                ),
                [],
            )
            if len(full_candidates) == 1:
                player = full_candidates[0]
            else:
                fallback_candidates = by_name_role.get(
                    (
                        _restore_key(row.get("Giocatore")),
                        _restore_key(row.get("Ruolo")),
                    ),
                    [],
                )
                if len(fallback_candidates) == 1:
                    player = fallback_candidates[0]
        if player is None:
            unmatched.append(_cell_text(row.get("Giocatore")))
            continue
        clean_id = str(player.get("id") or "")
        if clean_id in used_player_ids:
            player_name = _cell_text(row.get("Giocatore"))
            raise ValueError(f"Il giocatore {player_name} compare piu volte.")
        used_player_ids.add(clean_id)
        matched.append({**row, "_player": player})
    return matched, unmatched


def _tier_definitions(
    workbook: Any,
    matched_rows: list[dict[str, Any]],
    league: dict[str, Any],
) -> list[tuple[str, str]]:
    definitions: list[tuple[str, str]] = []
    seen: set[str] = set()
    if "Fasce personali" in workbook.sheetnames:
        sheet = workbook["Fasce personali"]
        for values in sheet.iter_rows(min_row=2, values_only=True):
            name = _cell_text(values[0] if values else None)
            if not name or _restore_key(name) in seen:
                continue
            color = _cell_text(values[1] if len(values) > 1 else "gray").lower()
            definitions.append((name, color if color in TIER_FILLS else "gray"))
            seen.add(_restore_key(name))
    if not definitions:
        for tier in league.get("auction_tiers", []):
            name = _cell_text(tier.get("name"))
            if not name or _restore_key(name) in seen:
                continue
            color = _cell_text(tier.get("color")).lower()
            definitions.append((name, color if color in TIER_FILLS else "gray"))
            seen.add(_restore_key(name))
    for row in matched_rows:
        name = _cell_text(row.get("Fascia personale"))
        if name and _restore_key(name) not in seen:
            definitions.append((name, "gray"))
            seen.add(_restore_key(name))
    return definitions


def _restore_auction_rows(
    league: dict[str, Any],
    rows: list[dict[str, Any]],
    tier_ids: dict[str, str],
) -> int:
    managers = auction_managers(league)
    if not managers:
        raise ValueError("Aggiungi almeno due partecipanti prima di importare l'asta.")
    league["purchases"] = []
    for manager in managers:
        manager["purchases"] = []
    league["auction_sale_events"] = []
    league["auction_history"] = []

    manager_flags: dict[str, bool] = {}
    manager_labels: dict[str, str] = {}
    for row in rows:
        name = _cell_text(row.get("Fantaallenatore"))
        if not name:
            if _parse_credits(row.get("Crediti")) is not None:
                player_name = _cell_text(row.get("Giocatore"))
                raise ValueError(
                    f"Crediti presenti senza fantaallenatore per {player_name}."
                )
            continue
        key = _restore_key(name)
        is_user = _is_yes(row.get("In rosa"))
        if key in manager_flags and manager_flags[key] != is_user:
            raise ValueError(f"Il fantaallenatore {name} ha assegnazioni incoerenti.")
        manager_flags[key] = is_user
        manager_labels[key] = name

    user_keys = [key for key, is_user in manager_flags.items() if is_user]
    if len(user_keys) > 1:
        raise ValueError("Nel file risultano piu nomi associati alla tua squadra.")
    user_manager = next(manager for manager in managers if manager.get("is_user"))
    manager_ids: dict[str, str] = {}
    if user_keys:
        user_key = user_keys[0]
        user_manager["name"] = manager_labels[user_key]
        manager_ids[user_key] = str(user_manager["id"])

    available_opponents = [
        manager for manager in managers if not manager.get("is_user")
    ]
    used_opponent_ids: set[str] = set()
    opponent_keys = [key for key, is_user in manager_flags.items() if not is_user]
    for key in opponent_keys:
        matching = next(
            (
                manager
                for manager in available_opponents
                if str(manager.get("id")) not in used_opponent_ids
                and _restore_key(manager.get("name")) == key
            ),
            None,
        )
        if matching is None:
            matching = next(
                (
                    manager
                    for manager in available_opponents
                    if str(manager.get("id")) not in used_opponent_ids
                ),
                None,
            )
        if matching is None:
            raise ValueError(
                "Il file contiene piu avversari della lega selezionata. "
                "Aumenta prima il numero di partecipanti nelle impostazioni."
            )
        matching["name"] = manager_labels[key]
        manager_ids[key] = str(matching["id"])
        used_opponent_ids.add(str(matching["id"]))

    changes: list[dict[str, Any]] = []
    purchases = 0
    for row in rows:
        manager_name = _cell_text(row.get("Fantaallenatore"))
        tier_name = _cell_text(row.get("Fascia personale"))
        note = _cell_text(row.get("Note personali"))
        change: dict[str, Any] = {"player": row["_player"]}
        if manager_name:
            credits = _parse_credits(row.get("Crediti"))
            if credits is None or credits <= 0:
                raise ValueError(
                    f"Crediti non validi per {_cell_text(row.get('Giocatore'))}."
                )
            change.update(
                {
                    "manager_id": manager_ids[_restore_key(manager_name)],
                    "price": credits,
                    "update_assignment": True,
                }
            )
            purchases += 1
        else:
            change["update_assignment"] = False
        if tier_name:
            change["tier_id"] = tier_ids[_restore_key(tier_name)]
        if "Note personali" in row:
            change["note"] = note
        if manager_name or tier_name or "Note personali" in row:
            changes.append(change)
    update_auction_assignments(league, changes)
    return purchases


def _restore_list_rows(
    league: dict[str, Any],
    rows: list[dict[str, Any]],
    tier_ids: dict[str, str],
) -> int:
    league["purchases"] = []
    changes = []
    purchases = 0
    for row in rows:
        in_roster = _is_yes(row.get("In rosa"))
        tier_name = _cell_text(row.get("Fascia personale"))
        change: dict[str, Any] = {
            "player": row["_player"],
            "in_roster": in_roster,
        }
        if tier_name:
            change["tier_id"] = tier_ids[_restore_key(tier_name)]
        if "Note personali" in row:
            change["note"] = _cell_text(row.get("Note personali"))
        changes.append(change)
        purchases += int(in_roster)
    update_list_assignments(league, changes)
    return purchases


def _parse_credits(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Valore crediti non valido: {value}.") from exc
    if not math.isfinite(result):
        raise ValueError("Il valore dei crediti deve essere un numero finito.")
    return result


def _is_yes(value: Any) -> bool:
    return _restore_key(value) in {"si", "yes", "true", "1", "x"}


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _restore_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _cell_text(value))
    ascii_text = text.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_text.casefold().split())


def _style_listone_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="173B33")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "Giocatore": 24,
        "Squadra": 12,
        "Ruolo": 9,
        "Fascia personale": 22,
        "Note personali": 42,
        "In rosa": 11,
        "Fantaallenatore": 22,
        "Crediti": 11,
    }
    for index, header in enumerate((cell.value for cell in sheet[1]), start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = widths.get(
            str(header), 16
        )
        if header in {"Quotazione", "FVM / 1000", "Crediti"}:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value_cell in cell:
                    value_cell.number_format = "#,##0"
        elif header in {"FM attesa", "Gol attesi", "Assist attesi"}:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value_cell in cell:
                    value_cell.number_format = "0.00"
        elif header in {
            "Titolarita %",
            "Affidabilita",
            "Rischio",
            "Propensione bonus",
            "Potenziale",
            "Indice",
        }:
            for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                for value_cell in cell:
                    value_cell.number_format = "0.0"


def _style_tiers_sheet(sheet: Any) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="173B33")
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 22


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if math.isnan(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _percentage_metric(value: Any) -> float | None:
    if value is None:
        return None
    try:
        metric = float(value)
    except (TypeError, ValueError):
        return None
    if 0 < metric <= 1:
        metric *= 100
    return round(max(0.0, min(100.0, metric)), 1)
